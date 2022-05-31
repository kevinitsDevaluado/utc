from itertools import count
import json
from datetime import datetime
from re import A
from django.contrib import messages
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.template.loader import render_to_string
import smtplib
from django.http import JsonResponse, HttpResponse,HttpResponseRedirect

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, FloatField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import requires_csrf_token
from django.views.generic import TemplateView
from django.views.generic import ListView, CreateView, UpdateView, DeleteView,TemplateView
from django.urls import reverse_lazy
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook

from core.barrios.forms import *
from core.barrios.models import *

class VerDistrictView(LoginRequiredMixin,CreateView):
    template_name = 'view_district/district.html'
    form_class = DistrictForm


    def get_success_url(self):
        pk = self.kwargs['pk']
        return reverse_lazy('verBarrio', kwargs={'pk': pk})

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()
            pk = self.kwargs['pk']
            fecha = datetime.now().strftime('%Y-%m-%d')
            if type == 'fecha':
                if Client.objects.filter(code__icontains=obj):
                    data['valid'] = False
            elif type == 'code':
                if RollCall.objects.filter(user__code__icontains=obj,date=fecha):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def send_email(self, id,subject):
        url = settings.LOCALHOST if not settings.DEBUG else self.request.META['HTTP_HOST']
        user = Client.objects.get(id=id)
        tema = Meeting.objects.get(subject__icontains=subject)
        message = MIMEMultipart('alternative')
        message['Subject'] = subject
        message['From'] = settings.EMAIL_HOST_USER
        message['To'] = user.user.email

        parameters = {
            'user': user,
            'tema': tema,
            'mainpage': Mainpage.objects.first(),
            'link_home': 'http://{}'.format(url),
            'link_login': 'http://{}/login'.format(url),
        }

        html = render_to_string('view_district/email_sign_in.html', parameters)
        content = MIMEText(html, 'html')
        message.attach(content)
        server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
        server.starttls()
        server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
        server.sendmail(
            settings.EMAIL_HOST_USER, user.user.email, message.as_string()
        )
        server.quit()


    def send_email_fees(self, id,subject):
            url = settings.LOCALHOST if not settings.DEBUG else self.request.META['HTTP_HOST']
            user = Client.objects.get(id=id)
            tema = Meeting.objects.get(subject__icontains=subject)
            message = MIMEMultipart('alternative')
            message['Subject'] = subject
            message['From'] = settings.EMAIL_HOST_USER
            message['To'] = user.user.email

            parameters = {
                'user': user,
                'tema': tema,
                'mainpage': Mainpage.objects.first(),
                'link_home': 'http://{}'.format(url),
                'link_login': 'http://{}/login'.format(url),
            }

            html = render_to_string('view_district/email_sign_in.html', parameters)
            content = MIMEText(html, 'html')
            message.attach(content)
            server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            server.starttls()
            server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
            server.sendmail(
                settings.EMAIL_HOST_USER, user.user.email, message.as_string()
            )
            server.quit()



    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'add_president':
                with transaction.atomic():
                    actualizar = PresidentDistrict.objects.all()
                    for l in actualizar:
                        l.state = False
                        l.save()

                    pk = self.kwargs['pk']
                    print('PRESIDENTES')
                    periodo = request.POST['period']
                    print('PERIODO',periodo)
                    user = request.POST['user']
                    print('PRESIDENTE ELEGIDO',user)
                    crear = PresidentDistrict()
                    crear.district_id = pk
                    crear.user_id = user
                    crear.period = periodo
                    crear.save()
                    messages.info(request, 'Se ha actualizado el Presidente..')


                    

            elif action == "list_client":
                with transaction.atomic():
                    print("que onda perritos")
                    code = request.POST['code']
                    fecha = datetime.now().strftime('%Y-%m-%d')
                    id_barrio = self.kwargs['pk']
                    if Client.objects.filter(code__icontains=code).exists():
                        print("codigo vale: ", code)
                        cliente = Client.objects.get(code__icontains=code)
                        if RollCall.objects.filter(user_id = cliente.id).exists() and RollCall.objects.filter(date = fecha).exists():
                            messages.info(request, 'La asistencia ya ha sido registrada..')                        
                        else:
                            a = RollCall()
                            a.user_id = cliente.id
                            a.date = fecha
                            a.district_id = id_barrio
                            a.save()
                            messages.info(request, 'Se ha registrado la asistencia correctamente..')
                    
                    else:
                        messages.info(request, 'El código de asistencia es incorrecto..')
            elif action == 'add_cuota':
                with transaction.atomic():
                    print('AGREGAR CUOTA')
                    id_barrio = self.kwargs['pk']
                    cuota = request.POST['cuota']
                    monto = request.POST['monto']
                    k = Fees()
                    k.district_id = id_barrio
                    k.cuota = cuota
                    k.monto = monto
                    k.save()
                    usuarios = Client.objects.filter(district_id = id_barrio)
                    for usu in usuarios:
                        deuda = Debt()
                        deuda.fees = k
                        deuda.monto = monto
                        deuda.user_id = usu.id
                        deuda.save()
                    
                    messages.info(request, 'Cuota Creada..')
                
            elif action == 'add_meeting':
                with transaction.atomic():
                    print('AGREGAR REUNIÓN')
                    id_barrio = self.kwargs['pk']
                    subject = request.POST['subject']
                    description = request.POST['description']
                    fecha_meeting = request.POST['fecha']

                    print('ASUNTO',subject)
                    print('DESCRIPCION',description)
                    print('FECHA',fecha_meeting)
                    m = Meeting()
                    m.subject = subject
                    m.description = description
                    m.date = fecha_meeting
                    m.district_id = id_barrio
                    m.save()
                    #clientes del barrio
                    m = Client.objects.filter(district_id = id_barrio)
                    for co in m:
                        self.send_email(co.id,subject)
                    
                    messages.info(request, 'Reunión Creada con exito..')

            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')


    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        id_barrio = self.kwargs['pk']
        barrio = District.objects.get(id= id_barrio)
        presidentes = PresidentDistrict.objects.filter(district_id = id_barrio)
        cuotas = Fees.objects.filter(district_id = id_barrio)
        fecha = datetime.now().strftime('%Y-%m-%d')
        meeting = Meeting.objects.filter(district_id = id_barrio)
        usuarios = Client.objects.filter(district_id = id_barrio)
        context['title'] = 'Solamente un pueblo culto puede ser veramente libre.'
        context['titulo'] = 'Matrículas'
        context['barrio'] = barrio
        context['cuota'] = cuotas
        context['fecha'] = fecha
        context['reuniones'] = meeting
        context['formPresiden'] = PresidentDistrictForm()
        context['presidentes'] = presidentes
        context['user'] = usuarios
        context['list_url'] = reverse_lazy('verBarrio', kwargs={'pk': self.kwargs['pk']})
        return context


class VerMeetingView(LoginRequiredMixin,CreateView):
    template_name = 'view_district/meeting.html'
    form_class = DistrictForm


    def get_success_url(self):
        pk = self.kwargs['pk']
        return reverse_lazy('verMeeting', kwargs={'pk': pk})

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()
            pk = self.kwargs['pk']
            fecha = datetime.now().strftime('%Y-%m-%d')
            if type == 'fecha':
                if Client.objects.filter(code__icontains=obj):
                    data['valid'] = False
            elif type == 'code':
                if RollCall.objects.filter(user__code__icontains=obj,date=fecha):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def send_email(self, id,subject,meeting):
        url = settings.LOCALHOST if not settings.DEBUG else self.request.META['HTTP_HOST']
        user = Client.objects.get(id=id)
        me = Meeting.objects.get(id=meeting)
        message = MIMEMultipart('alternative')
        message['Subject'] = subject
        message['From'] = settings.EMAIL_HOST_USER
        message['To'] = user.user.email

        parameters = {
            'user': user,
            'tema':me,
            'mainpage': Mainpage.objects.first(),
            'link_home': 'http://{}'.format(url),
            'link_login': 'http://{}/login'.format(url),
        }

        html = render_to_string('view_district/asisencia_email.html', parameters)
        content = MIMEText(html, 'html')
        message.attach(content)
        server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
        server.starttls()
        server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
        server.sendmail(
            settings.EMAIL_HOST_USER, user.user.email, message.as_string()
        )
        server.quit()

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'add_president':
                with transaction.atomic():
                    pass      
            elif action == "list_client":
                with transaction.atomic():
                    print("que onda perritos")
                    code = request.POST['code']
                    fecha = datetime.now().strftime('%Y-%m-%d')
                    id_meeting = self.kwargs['pk']
                    if Client.objects.filter(code__icontains=code).exists():
                        print("codigo vale: ", code)
                        cliente = Client.objects.get(code__icontains=code)
                        if RollCall.objects.filter(user_id = cliente.id).exists() and RollCall.objects.filter(meeting_id = id_meeting).exists():
                            messages.info(request, 'La asistencia ya ha sido registrada..')                        
                        else:
                            a = RollCall()
                            a.user_id = cliente.id
                            a.meeting_id = id_meeting
                            a.save()
                            self.send_email(cliente.id,"Confirmación de Asistencia",id_meeting)
                            messages.info(request, 'La asistencia Registrada de : {} {} '.format(cliente.user.first_name, cliente.user.last_name))
                    
                    else:
                        messages.info(request, 'El código de asistencia es incorrecto..')
           
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')





    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        id_meeting = self.kwargs['pk']
        info = Meeting.objects.get(id=id_meeting)
        countClient = Client.objects.filter(district_id = info.district ).count()
        barrio = District.objects.get(id = info.district.id)
        countParticipantes = RollCall.objects.filter(meeting_id = id_meeting).count()
        names = RollCall.objects.filter(meeting_id = id_meeting)
        context['title'] = 'Solamente un pueblo culto puede ser veramente libre.'
        context['titulo'] = 'Matrículas'
        context['info'] = info
        context['barrio'] = barrio
        context['count'] = countClient
        context['asistentes'] = countParticipantes
        context['names'] = names
        context['id_meeting'] = id_meeting
        context['list_url'] = reverse_lazy('verMeeting', kwargs={'pk': self.kwargs['pk']})
        return context


class VerFeesView(LoginRequiredMixin,CreateView):
    template_name = 'view_district/cuotas.html'
    form_class = DistrictForm

    
    def get_success_url(self):
        pk = self.kwargs['pk']
        return reverse_lazy('verFees', kwargs={'pk': pk})

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()

            fecha = datetime.now().strftime('%Y-%m-%d')
            if type == 'fecha':
                if Client.objects.filter(code__icontains=obj):
                    data['valid'] = False
            elif type == 'code':
                if RollCall.objects.filter(user__code__icontains=obj,date=fecha):
                    data['valid'] = False
            elif type == 'code  ':
                if RollCall.objects.filter(user__code__icontains=obj,date=fecha):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def send_email(self, id,subject,meeting):
        url = settings.LOCALHOST if not settings.DEBUG else self.request.META['HTTP_HOST']
        user = Client.objects.get(id=id)
        me = Meeting.objects.get(id=meeting)
        message = MIMEMultipart('alternative')
        message['Subject'] = subject
        message['From'] = settings.EMAIL_HOST_USER
        message['To'] = user.user.email

        parameters = {
            'user': user,
            'tema':me,
            'mainpage': Mainpage.objects.first(),
            'link_home': 'http://{}'.format(url),
            'link_login': 'http://{}/login'.format(url),
        }

        html = render_to_string('view_district/asisencia_email.html', parameters)
        content = MIMEText(html, 'html')
        message.attach(content)
        server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
        server.starttls()
        server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
        server.sendmail(
            settings.EMAIL_HOST_USER, user.user.email, message.as_string()
        )
        server.quit()

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'add_pagar':
                with transaction.atomic():
                    pass
            elif action == "list_payment":
                with transaction.atomic():
                    pass
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        id_fees = self.kwargs['pk']
        info = Fees.objects.get(id = id_fees)
        barrio = District.objects.get(id = info.district.id)
        cdebt = Debt.objects.filter(fees_id = id_fees, state = True).count()
        cdebtp = Debt.objects.filter(fees_id = id_fees, state = False).count()
        debt = Debt.objects.filter(fees_id = id_fees)
        context['title'] = 'Solamente un pueblo culto puede ser veramente libre.'
        context['titulo'] = 'Matrículas'
        context['info'] = info
        context['count'] = cdebt
        context['count2'] = cdebtp
        context['debt'] = debt
        context['barrio'] = barrio
        context['list_url'] = reverse_lazy('verFees', kwargs={'pk': self.kwargs['pk']})
        return context


class VerPaysmentsView(LoginRequiredMixin,CreateView):
    template_name = 'view_district/payments.html'
    form_class = DistrictForm

    def get_success_url(self):
        pk = self.kwargs['pk']
        return reverse_lazy('verPagos', kwargs={'pk': pk})

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()

            fecha = datetime.now().strftime('%Y-%m-%d')
            if type == 'fecha':
                if Client.objects.filter(code__icontains=obj):
                    data['valid'] = False
            elif type == 'code':
                if RollCall.objects.filter(user__code__icontains=obj,date=fecha):
                    data['valid'] = False
            elif type == 'code  ':
                if RollCall.objects.filter(user__code__icontains=obj,date=fecha):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def send_email(self, id,subject,debt,pago):
        url = settings.LOCALHOST if not settings.DEBUG else self.request.META['HTTP_HOST']
        user = Client.objects.get(id=id)
        deb = Debt.objects.get(id=debt)
        pay = Payments.objects.filter(debt_id = deb.id,user_id = deb.user.id)
        
        message = MIMEMultipart('alternative')
        message['Subject'] = subject
        message['From'] = settings.EMAIL_HOST_USER
        message['To'] = user.user.email

        parameters = {
            'user': user,
            'deb':deb,
            'pay':pay,
            'pago':pago,
            'mainpage': Mainpage.objects.first(),
            'link_home': 'http://{}'.format(url),
            'link_login': 'http://{}/login'.format(url),
        }

        html = render_to_string('view_district/pago_email.html', parameters)
        content = MIMEText(html, 'html')
        message.attach(content)
        server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
        server.starttls()
        server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
        server.sendmail(
            settings.EMAIL_HOST_USER, user.user.email, message.as_string()
        )
        server.quit()

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'add_pagar':
                with transaction.atomic():
                    pk = self.kwargs['pk']
                    valor = float(request.POST['valor'])
                    deb = Debt.objects.get(id=pk)
                    pa = Payments()
                    pa.debt_id = pk
                    pa.monto = valor
                    pa.user_id = deb.user.id
                    pa.save()

                    anterior = deb.monto
                    actual = float(anterior) - valor
                    deb.monto = actual
                    if actual == 0:
                        deb.state = False
                    else:
                        deb.state = True
                    deb.save()
                    self.send_email(deb.user.id,"Pago Efectuado",pk,valor)
                    messages.info(request, 'Pago efectuado con exito..')

                    

            elif action == "list_payment":
                with transaction.atomic():
                    pass
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        pk = self.kwargs['pk']
        deb = Debt.objects.get(id=pk)
        pay = Payments.objects.filter(debt_id = deb.id,user_id = deb.user.id)
        context['title'] = 'Solamente un pueblo culto puede ser veramente libre.'
        context['titulo'] = 'Matrículas'
        context['deb'] = deb
        context['pay'] = pay
        context['list_url'] = reverse_lazy('verPagos', kwargs={'pk': self.kwargs['pk']})
        return context


class report(LoginRequiredMixin,TemplateView):
    
    def get(self, request, *args, **kwargs):
        id_meeting = self.kwargs['pk']
        print('id barrio: ',id_meeting)
        names = RollCall.objects.filter(meeting_id = id_meeting)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE ASISTENTES'

        ws.merge_cells('B1:E1')

        ws['B3'] = 'NOMBRES'
        ws['C3'] = 'APELLIDOS'
        ws['D3'] = 'CÉDULA'
        ws['E3'] = 'CÓDIGO DE ASISTENCIA'

        cont = 4

        for me in names:
            ws.cell(row = cont,column = 2).value = me.user.user.first_name
            ws.cell(row = cont,column = 3).value = me.user.user.last_name
            ws.cell(row = cont,column = 4).value = me.user.user.dni
            ws.cell(row = cont,column = 5).value = me.user.code
            cont+=1

        nombre_archivo = "ReporteDeAsistencia.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment;  filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
